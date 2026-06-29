"""XLSX -> Markdown converter, backed by openpyxl.

Each worksheet becomes a ``## Sheet: <name>`` section, with its rows rendered as
a GitHub-flavoured Markdown table (the first row treated as the header).
"""

from __future__ import annotations

from openpyxl import load_workbook

from .base import BaseConverter


def _cell(value):
    """Render a single cell value as Markdown-safe text."""
    if value is None:
        return ""
    return str(value).replace("|", "\\|")


def _rows_to_markdown(rows):
    """Render a list of row tuples as a Markdown table, or ``""`` if empty."""
    if not rows:
        return ""

    width = max(len(r) for r in rows)
    header = [_cell(v) for v in rows[0]]
    header += [""] * (width - len(header))

    lines = [
        "| " + " | ".join(header) + " |",
        "| " + " | ".join("---" for _ in range(width)) + " |",
    ]
    for row in rows[1:]:
        cells = [_cell(v) for v in row]
        cells += [""] * (width - len(cells))
        lines.append("| " + " | ".join(cells) + " |")
    return "\n".join(lines)


class XLSXConverter(BaseConverter):
    """Convert a .xlsx workbook into Markdown, one section per sheet."""

    extensions = (".xlsx",)

    def _convert(self, file_path: str) -> str:
        workbook = load_workbook(file_path, data_only=True)
        sections = []

        for sheet in workbook.worksheets:
            # Drop trailing fully-empty rows so blank sheets are skipped.
            rows = [
                row
                for row in sheet.iter_rows(values_only=True)
                if any(cell is not None for cell in row)
            ]
            if not rows:
                continue

            table = _rows_to_markdown(rows)
            sections.append(f"## Sheet: {sheet.title}\n\n{table}")

        return "\n\n".join(sections)
