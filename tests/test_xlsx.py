"""Tests for XLSXConverter. Fixtures are generated with openpyxl at runtime."""

from openpyxl import Workbook

from pydown.converters.xlsx import XLSXConverter


def test_single_sheet_as_table(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "People"
    ws.append(["Name", "Age"])
    ws.append(["Alice", 30])
    ws.append(["Bob", 25])
    path = tmp_path / "book.xlsx"
    wb.save(str(path))

    md = XLSXConverter().convert(str(path))

    assert "## Sheet: People" in md
    assert "| Name | Age |" in md
    assert "| --- | --- |" in md
    assert "| Alice | 30 |" in md
    assert "| Bob | 25 |" in md


def test_multiple_sheets(tmp_path):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "First"
    ws1.append(["a", "b"])
    ws2 = wb.create_sheet("Second")
    ws2.append(["c", "d"])
    path = tmp_path / "multi.xlsx"
    wb.save(str(path))

    md = XLSXConverter().convert(str(path))

    assert "## Sheet: First" in md
    assert "## Sheet: Second" in md
    assert md.index("## Sheet: First") < md.index("## Sheet: Second")


def test_empty_sheet_is_skipped(tmp_path):
    wb = Workbook()
    wb.active.title = "Blank"  # left empty
    data = wb.create_sheet("Data")
    data.append(["x"])
    data.append([1])
    path = tmp_path / "withblank.xlsx"
    wb.save(str(path))

    md = XLSXConverter().convert(str(path))

    assert "## Sheet: Blank" not in md
    assert "## Sheet: Data" in md
